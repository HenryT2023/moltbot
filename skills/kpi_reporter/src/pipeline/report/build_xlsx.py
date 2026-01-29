#!/usr/bin/env python3
"""
Excel 报表生成脚本
读取 JSON 数据，生成包含 Summary/BySource/RawPreview 三个 Sheet 的 XLSX 文件
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError as e:
    print(f"缺少依赖: {e}", file=sys.stderr)
    print("请运行: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def create_xlsx(data: dict, output_path: str) -> None:
    """生成 Excel 报表"""
    wb = Workbook()

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # 标题
    ws_summary["A1"] = f"KPI 报表 - {data['timeWindow']['label']}"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.merge_cells("A1:D1")

    ws_summary["A2"] = f"时间范围: {data['timeWindow']['start']} ~ {data['timeWindow']['end']}"
    ws_summary.merge_cells("A2:D2")

    # KPI 表头
    headers = ["指标", "当前值", "上期值", "环比变化"]
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # KPI 数据
    current_kpis = {k["name"]: k["value"] for k in data["kpis"]["current"]}
    previous_kpis = {k["name"]: k["value"] for k in data["kpis"]["previous"]}

    row = 5
    for kpi in data["kpis"]["current"]:
        name = kpi["name"]
        current = kpi["value"]
        previous = previous_kpis.get(name, 0)

        # 计算环比
        if previous != 0:
            change = ((current - previous) / previous) * 100
            change_str = f"{change:+.1f}%"
        else:
            change_str = "N/A"

        # 格式化数值
        if current == int(current):
            current_str = f"{int(current):,}"
        else:
            current_str = f"{current:,.2f}"

        if previous == int(previous):
            previous_str = f"{int(previous):,}"
        else:
            previous_str = f"{previous:,.2f}"

        ws_summary.cell(row=row, column=1, value=name).border = thin_border
        ws_summary.cell(row=row, column=2, value=current_str).border = thin_border
        ws_summary.cell(row=row, column=3, value=previous_str).border = thin_border
        ws_summary.cell(row=row, column=4, value=change_str).border = thin_border
        row += 1

    # 调整列宽
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 15
    ws_summary.column_dimensions["C"].width = 15
    ws_summary.column_dimensions["D"].width = 15

    # Sheet 2: BySource
    ws_source = wb.create_sheet("BySource")

    ws_source["A1"] = "按数据源分类"
    ws_source["A1"].font = Font(bold=True, size=14)
    ws_source.merge_cells("A1:C1")

    row = 3
    for source, metrics in data["bySource"].items():
        ws_source.cell(row=row, column=1, value=source.upper())
        ws_source.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        # 表头
        ws_source.cell(row=row, column=1, value="指标").font = header_font
        ws_source.cell(row=row, column=1).fill = header_fill
        ws_source.cell(row=row, column=2, value="数值").font = header_font
        ws_source.cell(row=row, column=2).fill = header_fill
        row += 1

        for m in metrics:
            ws_source.cell(row=row, column=1, value=m["metric"])
            value = m["value"]
            if value == int(value):
                ws_source.cell(row=row, column=2, value=f"{int(value):,}")
            else:
                ws_source.cell(row=row, column=2, value=f"{value:,.2f}")
            row += 1

        row += 1  # 空行分隔

    ws_source.column_dimensions["A"].width = 20
    ws_source.column_dimensions["B"].width = 15

    # Sheet 3: RawPreview
    ws_raw = wb.create_sheet("RawPreview")

    ws_raw["A1"] = "原始数据预览（前 100 条）"
    ws_raw["A1"].font = Font(bold=True, size=14)
    ws_raw.merge_cells("A1:E1")

    if data["rawPreview"]:
        # 表头
        raw_headers = ["日期", "来源", "指标", "数值", "维度"]
        for col, header in enumerate(raw_headers, 1):
            cell = ws_raw.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 数据
        row = 4
        for record in data["rawPreview"]:
            ws_raw.cell(row=row, column=1, value=record["date"])
            ws_raw.cell(row=row, column=2, value=record["source"])
            ws_raw.cell(row=row, column=3, value=record["metric"])
            ws_raw.cell(row=row, column=4, value=record["value"])
            dims_str = json.dumps(record.get("dims", {}), ensure_ascii=False)
            ws_raw.cell(row=row, column=5, value=dims_str if dims_str != "{}" else "")
            row += 1

    ws_raw.column_dimensions["A"].width = 12
    ws_raw.column_dimensions["B"].width = 10
    ws_raw.column_dimensions["C"].width = 15
    ws_raw.column_dimensions["D"].width = 12
    ws_raw.column_dimensions["E"].width = 40

    # 保存
    wb.save(output_path)
    print(f"报表已生成: {output_path}")


def main():
    if len(sys.argv) != 3:
        print("用法: python build_xlsx.py <input_json> <output_xlsx>", file=sys.stderr)
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2]

    try:
        with open(input_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        create_xlsx(data, output_path)
    except Exception as e:
        print(f"生成报表失败: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
